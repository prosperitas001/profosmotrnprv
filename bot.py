import os
import io
import base64
import logging
from datetime import datetime

import anthropic
from telegram import Update
from telegram.ext import ApplicationBuilder, CommandHandler, MessageHandler, filters, ContextTypes
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
import json

logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

TELEGRAM_TOKEN = os.environ["TELEGRAM_TOKEN"]
ANTHROPIC_API_KEY = os.environ["ANTHROPIC_API_KEY"]

sessions: dict[int, list[dict]] = {}

anthropic_client = anthropic.Anthropic(api_key=ANTHROPIC_API_KEY)


def build_excel(rows: list[dict]) -> bytes:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Профосмотр"

    headers = ["ФИО", "Дата рождения", "Пол", "Профессия"]
    widths = [35, 16, 8, 22]

    header_fill = PatternFill("solid", fgColor="D9E1F2")
    header_font = Font(name="Arial", bold=True, size=10)
    cell_font = Font(name="Arial", size=10)
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left = Alignment(horizontal="left", vertical="center", wrap_text=True)
    thin = Side(style="thin", color="AAAAAA")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for col, (h, w) in enumerate(zip(headers, widths), 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center
        cell.border = border
        ws.column_dimensions[cell.column_letter].width = w

    ws.row_dimensions[1].height = 28

    for row_idx, r in enumerate(rows, 2):
        for col_idx, key in enumerate(["fio", "dob", "gender", "profession"], 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=r.get(key, ""))
            cell.font = cell_font
            cell.border = border
            cell.alignment = center if col_idx in (2, 3) else left
        ws.row_dimensions[row_idx].height = 18

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def extract_fields(file_bytes: bytes, mime_type: str) -> dict:
    if mime_type == "application/pdf":
        content_block = {
            "type": "document",
            "source": {
                "type": "base64",
                "media_type": "application/pdf",
                "data": base64.b64encode(file_bytes).decode()
            }
        }
    else:
        content_block = {
            "type": "image",
            "source": {
                "type": "base64",
                "media_type": mime_type,
                "data": base64.b64encode(file_bytes).decode()
            }
        }

    response = anthropic_client.messages.create(
        model="claude-sonnet-4-6",
        max_tokens=500,
        messages=[{
            "role": "user",
            "content": [
                content_block,
                {
                    "type": "text",
                    "text": (
                        "Это направление на медицинский осмотр. "
                        "Извлеки ТОЛЬКО следующие поля и верни строго в формате JSON без markdown и пояснений:\n"
                        '{"fio": "ФИО из пункта 1", "dob": "дата рождения ДД.ММ.ГГГГ", '
                        '"gender": "муж или жен", "profession": "профессия из пункта 7"}\n'
                        "Если поле не найдено — пустая строка."
                    )
                }
            ]
        }]
    )

    text = "".join(b.text for b in response.content if hasattr(b, "text"))
    text = text.replace("```json", "").replace("```", "").strip()
    return json.loads(text)


# ─── Handlers ────────────────────────────────────────────────────────────────

async def cmd_start(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    await update.message.reply_text(
        "👋 Привет! Я бот для обработки направлений на профосмотр.\n\n"
        "📎 Отправьте фото или PDF направления — извлеку данные.\n"
        "📊 /excel — получить Excel-файл со всеми записями\n"
        "🗑 /clear — очистить список\n"
        "📋 /list — показать текущие записи"
    )


async def cmd_list(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    chat_id = update.effective_chat.id
    rows = sessions.get(chat_id, [])
    if not rows:
        await update.message.reply_text("Список пуст. Отправьте направление.")
        return
    text = f"📋 Записей: {len(rows)}\n\n"
    for i, r in enumerate(rows, 1):
        text += f"{i}. {r.get('fio','—')} | {r.get('dob','—')} | {r.get('gender','—')} | {r.get('profession','—')}\n"
    await update.message.reply_text(text)


async def cmd_clear(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    chat_id = update.effective_chat.id
    sessions[chat_id] = []
    await update.message.reply_text("🗑 Список очищен.")


async def cmd_excel(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    chat_id = update.effective_chat.id
    rows = sessions.get(chat_id, [])
    if not rows:
        await update.message.reply_text("Нет данных. Сначала отправьте направления.")
        return

    msg = await update.message.reply_text("⏳ Формирую Excel...")

    try:
        xlsx_bytes = build_excel(rows)
        date_str = datetime.now().strftime("%Y-%m-%d_%H-%M")
        filename = f"профосмотр_{date_str}.xlsx"

        await update.message.reply_document(
            document=io.BytesIO(xlsx_bytes),
            filename=filename,
            caption=f"✅ Готово! Записей: {len(rows)}"
        )
        await msg.delete()
    except Exception as e:
        logger.error(f"Excel error: {e}")
        await msg.edit_text(f"❌ Ошибка: {e}")


async def handle_photo(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    chat_id = update.effective_chat.id
    msg = await update.message.reply_text("🔍 Распознаю направление...")

    try:
        photo = update.message.photo[-1]
        file = await photo.get_file()
        file_bytes = bytes(await file.download_as_bytearray())
        fields = extract_fields(file_bytes, "image/jpeg")

        sessions.setdefault(chat_id, []).append(fields)

        await msg.edit_text(
            f"✅ Добавлено (всего: {len(sessions[chat_id])}):\n\n"
            f"👤 {fields.get('fio', '—')}\n"
            f"🎂 {fields.get('dob', '—')}\n"
            f"⚧ {fields.get('gender', '—')}\n"
            f"💼 {fields.get('profession', '—')}\n\n"
            f"Отправьте ещё или /excel для скачивания файла."
        )
    except Exception as e:
        logger.error(f"Photo error: {e}")
        await msg.edit_text(f"❌ Ошибка распознавания: {e}")


async def handle_document(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    chat_id = update.effective_chat.id
    doc = update.message.document

    if doc.mime_type not in ("application/pdf", "image/jpeg", "image/png"):
        await update.message.reply_text("⚠️ Поддерживаются только PDF, JPG, PNG.")
        return

    msg = await update.message.reply_text("🔍 Распознаю направление...")

    try:
        file = await doc.get_file()
        file_bytes = bytes(await file.download_as_bytearray())
        fields = extract_fields(file_bytes, doc.mime_type)

        sessions.setdefault(chat_id, []).append(fields)

        await msg.edit_text(
            f"✅ Добавлено (всего: {len(sessions[chat_id])}):\n\n"
            f"👤 {fields.get('fio', '—')}\n"
            f"🎂 {fields.get('dob', '—')}\n"
            f"⚧ {fields.get('gender', '—')}\n"
            f"💼 {fields.get('profession', '—')}\n\n"
            f"Отправьте ещё или /excel для скачивания файла."
        )
    except Exception as e:
        logger.error(f"Document error: {e}")
        await msg.edit_text(f"❌ Ошибка распознавания: {e}")


def main():
    app = ApplicationBuilder().token(TELEGRAM_TOKEN).build()
    app.add_handler(CommandHandler("start", cmd_start))
    app.add_handler(CommandHandler("list", cmd_list))
    app.add_handler(CommandHandler("clear", cmd_clear))
    app.add_handler(CommandHandler("excel", cmd_excel))
    app.add_handler(MessageHandler(filters.PHOTO, handle_photo))
    app.add_handler(MessageHandler(filters.Document.ALL, handle_document))
    logger.info("Bot started")
    app.run_polling()


if __name__ == "__main__":
    main()
